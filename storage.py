import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

DATA_DIR = Path(__file__).parent / "data"
LOGS_DIR = DATA_DIR / "logs"
META_FILE = DATA_DIR / "meta.json"
SUBJECTS_FILE = DATA_DIR / "subjects.json"
TIMETABLE_FILE = DATA_DIR / "timetable.json"
MONTHLY_CSV_FILE = DATA_DIR / "attendance_monthly_summary.csv"
ALL_LOGS_CSV_FILE = DATA_DIR / "attendance_all_logs.csv"
EXCEL_FILE = DATA_DIR / "attendance_records.xlsx"

STORAGE_VERSION = 3


def _ensure_dirs() -> None:
    LOGS_DIR.mkdir(parents=True, exist_ok=True)


def _read_json(path: Path) -> Any | None:
    if not path.exists():
        return None
    with path.open(encoding="utf-8") as fh:
        return json.load(fh)


def _write_json(path: Path, payload: Any) -> None:
    _ensure_dirs()
    with path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)
        fh.write("\n")


def _weekday_name(date_str: str) -> str:
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    return dt.strftime("%A")


def _export_spreadsheets(state: dict) -> None:
    """Generate Excel (.xlsx) and CSV files for date-wise class records and month-wise subject breakdowns."""
    subjects = state.get("subjects", [])
    attendance_logs = state.get("attendanceLogs", {})

    subj_map = {s.get("id"): s for s in subjects if isinstance(s, dict) and s.get("id")}

    # Collect all unique months (YYYY-MM)
    months = sorted(list({d[:7] for d in attendance_logs.keys()}), reverse=True)
    if not months:
        months = [datetime.now().strftime("%Y-%m")]

    # 1. Prepare Date-wise Detailed Class Logs
    date_log_rows = []
    for dstr in sorted(attendance_logs.keys(), reverse=True):
        month_label = datetime.strptime(dstr[:7], "%Y-%m").strftime("%B %Y")
        wday = _weekday_name(dstr)
        day_logs = attendance_logs[dstr]
        for key, log in sorted(day_logs.items()):
            sid = log.get("subjectId")
            s_obj = subj_map.get(sid, {})
            date_log_rows.append({
                "Date": dstr,
                "Month": month_label,
                "Weekday": wday,
                "Time_Slot": log.get("timeSlot", ""),
                "Subject_Code": s_obj.get("code", ""),
                "Subject_Name": log.get("subjectName", s_obj.get("name", "")),
                "Class_Type": log.get("type", "Lecture"),
                "Status": log.get("status", "Unmarked"),
                "Log_Submitted": "Yes" if log.get("submitted") else ("No" if log.get("status") == "LoggedMissed" else "-"),
                "Note": log.get("note", ""),
                "Extra_Class": "Yes" if log.get("isExtra") else "No"
            })

    # Write Date-wise All Logs CSV
    with ALL_LOGS_CSV_FILE.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow([
            "Date", "Month", "Weekday", "Time Slot", "Subject Code",
            "Subject Name", "Class Type", "Attendance Status", "Log Submitted", "Event Note / Reason", "Extra Class"
        ])
        for r in date_log_rows:
            writer.writerow([
                r["Date"], r["Month"], r["Weekday"], r["Time_Slot"],
                r["Subject_Code"], r["Subject_Name"], r["Class_Type"],
                r["Status"], r["Log_Submitted"], r["Note"], r["Extra_Class"]
            ])

    # 2. Prepare Monthly Subject Breakdown Data (with Lecture, Tutorial, Practical counts)
    monthly_rows = []
    for month in months:
        month_label = datetime.strptime(month, "%Y-%m").strftime("%B %Y")
        for s in subjects:
            sid = s.get("id")
            sname = s.get("name", sid)
            scode = s.get("code", "")
            target = s.get("target", 67)

            by_type = {
                "Lecture":   {"attended": 0, "missed": 0, "logged": 0, "submitted": 0, "conducted": 0},
                "Tutorial":  {"attended": 0, "missed": 0, "logged": 0, "submitted": 0, "conducted": 0},
                "Practical": {"attended": 0, "missed": 0, "logged": 0, "submitted": 0, "conducted": 0}
            }

            tot_attended = 0
            tot_missed = 0
            tot_logged = 0
            tot_submitted = 0
            tot_cancelled = 0

            for dstr, day_logs in attendance_logs.items():
                if dstr.startswith(month):
                    for log in day_logs.values():
                        if log.get("subjectId") == sid:
                            st = log.get("status")
                            ctype = log.get("type", "Lecture")
                            if ctype not in by_type:
                                ctype = "Lecture"

                            if st == "Attended":
                                tot_attended += 1
                                by_type[ctype]["attended"] += 1
                                by_type[ctype]["conducted"] += 1
                            elif st == "Missed":
                                tot_missed += 1
                                by_type[ctype]["missed"] += 1
                                by_type[ctype]["conducted"] += 1
                            elif st == "LoggedMissed":
                                tot_logged += 1
                                by_type[ctype]["logged"] += 1
                                by_type[ctype]["conducted"] += 1
                                if log.get("submitted"):
                                    tot_submitted += 1
                                    by_type[ctype]["submitted"] += 1
                            elif st == "Cancelled":
                                tot_cancelled += 1

            tot_conducted = tot_attended + tot_missed + tot_logged
            phys_pct = round((tot_attended / tot_conducted * 100), 1) if tot_conducted > 0 else 100.0
            eff_pct = round(((tot_attended + tot_submitted) / tot_conducted * 100), 1) if tot_conducted > 0 else 100.0
            status_text = "On Track" if phys_pct >= target else ("Log Covered" if eff_pct >= target else "Critical")

            lec_str = f"{by_type['Lecture']['attended']}/{by_type['Lecture']['conducted']}" if by_type['Lecture']['conducted'] > 0 else "-"
            tut_str = f"{by_type['Tutorial']['attended']}/{by_type['Tutorial']['conducted']}" if by_type['Tutorial']['conducted'] > 0 else "-"
            prac_str = f"{by_type['Practical']['attended']}/{by_type['Practical']['conducted']}" if by_type['Practical']['conducted'] > 0 else "-"

            monthly_rows.append({
                "Month": month_label,
                "Subject_Code": scode,
                "Subject_Name": sname,
                "Target_%": target,
                "Lecture_Attended_Conducted": lec_str,
                "Tutorial_Attended_Conducted": tut_str,
                "Practical_Attended_Conducted": prac_str,
                "Total_Conducted": tot_conducted,
                "Total_Attended": tot_attended,
                "Missed": tot_missed,
                "Event_Logs": tot_logged,
                "Logs_Submitted": tot_submitted,
                "Cancelled": tot_cancelled,
                "Physical_Attendance_%": phys_pct,
                "Effective_Attendance_%": eff_pct,
                "Status": status_text
            })

    # Write Monthly Summary CSV
    with MONTHLY_CSV_FILE.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow([
            "Month", "Subject Code", "Subject Name", "Target %",
            "Lecture (Attended/Conducted)", "Tutorial (Attended/Conducted)", "Practical (Attended/Conducted)",
            "Total Conducted", "Total Attended", "Missed", "Event Logs", "Logs Submitted", "Cancelled",
            "Physical Attendance %", "Effective Attendance %", "Status"
        ])
        for r in monthly_rows:
            writer.writerow([
                r["Month"], r["Subject_Code"], r["Subject_Name"], r["Target_%"],
                r["Lecture_Attended_Conducted"], r["Tutorial_Attended_Conducted"], r["Practical_Attended_Conducted"],
                r["Total_Conducted"], r["Total_Attended"], r["Missed"], r["Event_Logs"],
                r["Logs_Submitted"], r["Cancelled"], r["Physical_Attendance_%"],
                r["Effective_Attendance_%"], r["Status"]
            ])

    # 3. Write Excel (.xlsx) if openpyxl is installed
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # Sheet 1: Date-wise Class Logs (Primary Sheet)
        ws1 = wb.active
        ws1.title = "Date-wise Class Records"
        ws1.views.sheetView[0].showGridLines = True

        headers1 = [
            "Date", "Month", "Weekday", "Time Slot", "Subject Code",
            "Subject Name", "Class Type", "Attendance Status", "Log Submitted", "Extra Class"
        ]

        ws1.append(headers1)
        for col_idx in range(1, len(headers1) + 1):
            cell = ws1.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r in date_log_rows:
            ws1.append([
                r["Date"], r["Month"], r["Weekday"], r["Time_Slot"],
                r["Subject_Code"], r["Subject_Name"], r["Class_Type"],
                r["Status"], r["Log_Submitted"], r["Extra_Class"]
            ])

        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=len(headers1)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        for col in ws1.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Sheet 2: Subject Month-wise Breakdown
        ws2 = wb.create_sheet(title="Subject Month Breakdown")
        ws2.views.sheetView[0].showGridLines = True

        headers2 = [
            "Month", "Subject Code", "Subject Name", "Target %",
            "Lecture (Att/Cond)", "Tutorial (Att/Cond)", "Practical (Att/Cond)",
            "Total Conducted", "Total Attended", "Missed", "Event Logs", "Submitted Logs", "Cancelled",
            "Physical %", "Effective %", "Status"
        ]
        ws2.append(headers2)
        for col_idx in range(1, len(headers2) + 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r in monthly_rows:
            ws2.append([
                r["Month"], r["Subject_Code"], r["Subject_Name"], r["Target_%"],
                r["Lecture_Attended_Conducted"], r["Tutorial_Attended_Conducted"], r["Practical_Attended_Conducted"],
                r["Total_Conducted"], r["Total_Attended"], r["Missed"], r["Event_Logs"],
                r["Logs_Submitted"], r["Cancelled"], r["Physical_Attendance_%"],
                r["Effective_Attendance_%"], r["Status"]
            ])

        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=len(headers2)):
            for cell in row:
                cell.border = thin_border

        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(EXCEL_FILE)

    except ImportError:
        pass
# openpyxl not installed, CSV files are generated natively
        pass


def state_to_files(state: dict) -> None:
    """Split app state into individual JSON files under data/ and export Excel/CSV."""
    _ensure_dirs()

    subjects = state.get("subjects", [])
    timetable = state.get("timetable", {})
    attendance_logs = state.get("attendanceLogs", {})

    _write_json(
        SUBJECTS_FILE,
        {
            "description": "All tracked subjects with attendance targets.",
            "subjects": subjects,
        },
    )

    _write_json(
        TIMETABLE_FILE,
        {
            "description": "Weekly class schedule. Keys are Mon–Sun.",
            "timetable": timetable,
        },
    )

    # Remove log files for dates no longer present in state.
    active_dates = set(attendance_logs.keys())
    for path in LOGS_DIR.glob("*.json"):
        if path.stem not in active_dates:
            path.unlink(missing_ok=True)

    for date_str, day_logs in sorted(attendance_logs.items()):
        entries = []
        for key, log in sorted(day_logs.items()):
            entries.append(
                {
                    "key": key,
                    "subjectId": log.get("subjectId"),
                    "subjectName": log.get("subjectName"),
                    "type": log.get("type"),
                    "status": log.get("status"),
                    "timeSlot": log.get("timeSlot"),
                    "isExtra": bool(log.get("isExtra", False)),
                    "submitted": log.get("submitted"),
                    "note": log.get("note", ""),
                }
            )

        _write_json(
            LOGS_DIR / f"{date_str}.json",
            {
                "date": date_str,
                "weekday": _weekday_name(date_str),
                "description": f"Attendance records for {date_str}.",
                "entries": entries,
            },
        )

    _write_json(
        META_FILE,
        {
            "version": STORAGE_VERSION,
            "lastUpdated": datetime.now(timezone.utc).isoformat(),
            "subjectCount": len(subjects),
            "logDayCount": len(attendance_logs),
        },
    )

    # Export spreadsheets (CSV & XLSX)
    try:
        _export_spreadsheets(state)
    except Exception as exc:
        print(f"AuraAttend: Warning, could not write spreadsheet exports: {exc}")


def files_to_state() -> dict | None:
    """Rebuild app state from data/ folder. Returns None if no saved data."""
    subjects_doc = _read_json(SUBJECTS_FILE)
    timetable_doc = _read_json(TIMETABLE_FILE)

    if subjects_doc is None and timetable_doc is None and not any(LOGS_DIR.glob("*.json")):
        return None

    subjects = (subjects_doc or {}).get("subjects", [])
    timetable = (timetable_doc or {}).get("timetable", {})

    attendance_logs: dict[str, dict] = {}
    for path in sorted(LOGS_DIR.glob("*.json")):
        day_doc = _read_json(path)
        if not day_doc:
            continue
        date_str = day_doc.get("date", path.stem)
        day_logs: dict[str, dict] = {}
        for entry in day_doc.get("entries", []):
            key = entry.get("key")
            if not key:
                continue
            day_logs[key] = {
                "subjectId": entry.get("subjectId"),
                "subjectName": entry.get("subjectName"),
                "type": entry.get("type"),
                "status": entry.get("status"),
                "timeSlot": entry.get("timeSlot"),
                "isExtra": entry.get("isExtra", False),
            }
            if entry.get("submitted") is not None:
                day_logs[key]["submitted"] = entry["submitted"]
            if entry.get("note"):
                day_logs[key]["note"] = entry["note"]
        if day_logs:
            attendance_logs[date_str] = day_logs

    return {
        "subjects": subjects,
        "timetable": timetable,
        "attendanceLogs": attendance_logs,
    }


def load_full_state() -> dict | None:
    return files_to_state()


def save_full_state(state: dict) -> None:
    state_to_files(state)


def get_last_updated() -> str | None:
    meta = _read_json(META_FILE)
    if not meta:
        return None
    return meta.get("lastUpdated")
